#!/usr/bin/env python3


import csv
import io
import json
import os
import queue
import re
import tempfile
import threading
import tkinter as tk
from tkinter import filedialog, ttk
from urllib.parse import urljoin

import requests
from requests.exceptions import RequestException



TOKEN_PATH = '/geri_connect/auth/v1/token'
PUBLIC_API_PREFIX = '/geri_connect/public/api/v1'
VIDEO_PATH = '/files/video.json/{session_uuid}/well{well_no:02d}_zid{zid}.mp4'
VIDEO_ZIDS = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '99']
DEFAULT_REQUEST_TIMEOUT = 30
DEFAULT_DOWNLOAD_TIMEOUT = 300
INPUT_COLUMN_ALIASES = {
    'patient_given_names': ['女方姓名'],
    'patient_name': ['男方姓名'],
    'no': ['no'],
}
OPTIONAL_INPUT_COLUMN_ALIASES = {
    'session_uuid': ['session_uuid'],
}


class CandlePublicClient:

    def __init__(self, base_url, username, password,
                 request_timeout=DEFAULT_REQUEST_TIMEOUT,
                 download_timeout=DEFAULT_DOWNLOAD_TIMEOUT):
        self.base_url = base_url.rstrip('/')
        self.request_timeout = request_timeout
        self.download_timeout = download_timeout
        self.s = requests.Session()
        self.s.headers.update(self._get_auth_header(username, password))

    def _get_auth_header(self, username, password):
        url = urljoin(self.base_url, TOKEN_PATH)
        try:
            response = requests.get(url,
                                    auth=(username, password),
                                    verify=False,
                                    timeout=self.request_timeout)
            response.raise_for_status()
            token = response.json().get('token')
        except RequestException as ex:
            raise RuntimeError('IP 不通或用户名密码错误: {}'.format(ex))
        except ValueError:
            raise RuntimeError('auth 接口返回内容不是 JSON。')

        if not token:
            raise RuntimeError('登录失败，请检查用户名和密码。')
        return {'Authorization': 'Bearer {}'.format(token)}

    def get_dishrecords(self):
        return self._get_public_json('dishrecords')

    def get_sessionrecords(self, dish_uuid):
        return self._get_public_json('sessionrecords', {'dish_uuid': dish_uuid})

    def get_sessionrecord(self, session_uuid):
        return self._get_public_json('sessionrecords', {'session_uuid': session_uuid})

    def _get_public_json(self, endpoint, params=None):
        url = urljoin(self.base_url, PUBLIC_API_PREFIX + '/' + endpoint)
        try:
            response = self.s.get(url, params=params, verify=False, timeout=self.request_timeout)
            response.raise_for_status()
            body = response.json()
        except RequestException as ex:
            raise RuntimeError('接口请求失败 {}: {}'.format(endpoint, ex))
        except ValueError:
            raise RuntimeError('接口返回内容不是 JSON: {}'.format(endpoint))

        if body.get('error'):
            raise RuntimeError(body['error'])
        return body.get('content') or []

    def download_video(self, session_uuid, well_no, zid, output_path, overwrite=False):
        if os.path.exists(output_path) and not overwrite:
            return 'skipped'

        url = urljoin(self.base_url, VIDEO_PATH.format(session_uuid=session_uuid, well_no=well_no, zid=zid))
        try:
            response = self.s.get(url, stream=True, verify=False, timeout=self.download_timeout)
            response.raise_for_status()
            with open(output_path, 'wb') as video_file:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        video_file.write(chunk)
        except RequestException as ex:
            raise RuntimeError(ex)
        return 'downloaded'


def normalize(value):
    return (value or '').strip().lower()


def make_patient_key(patient_given_names, patient_name):
    return normalize(patient_given_names), normalize(patient_name)


def safe_filename(value):
    value = '' if value is None else str(value).strip()
    value = re.sub(r'[<>:"/\\|?*\r\n\t]+', '_', value)
    return value.strip(' ._') or 'unknown'


def resolve_input_columns(fieldnames):
    fieldnames = [str(name).strip() if name is not None else '' for name in (fieldnames or [])]
    resolved = {}

    for canonical, aliases in INPUT_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in fieldnames:
                resolved[canonical] = alias
                break

    missing = [INPUT_COLUMN_ALIASES[key][0] for key in INPUT_COLUMN_ALIASES if key not in resolved]
    if missing:
        raise RuntimeError('Excel 缺少必需表头: {}'.format(', '.join(missing)))

    for canonical, aliases in OPTIONAL_INPUT_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in fieldnames:
                resolved[canonical] = alias
                break
    return resolved


def canonicalize_target(row, columns, row_number=None):
    return {
        'patient_given_names': row.get(columns['patient_given_names'], ''),
        'patient_name': row.get(columns['patient_name'], ''),
        'no': row.get(columns['no'], ''),
        'session_uuid': row.get(columns['session_uuid'], '') if 'session_uuid' in columns else '',
        '_row_number': row_number,
    }


def read_csv_targets(csv_path):
    encodings = ['utf-8-sig', 'gb18030']
    last_error = None

    for encoding in encodings:
        try:
            with io.open(csv_path, 'r', encoding=encoding, newline='') as csv_file:
                sample = csv_file.read(4096)
                csv_file.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
                except csv.Error:
                    dialect = csv.excel

                reader = csv.DictReader(csv_file, dialect=dialect)
                columns = resolve_input_columns(reader.fieldnames)
                return [canonicalize_target(row, columns, row_number=index + 2)
                        for index, row in enumerate(reader)]
        except UnicodeDecodeError as ex:
            last_error = ex

    if last_error:
        raise RuntimeError('CSV 编码无法识别，请保存为 UTF-8 或 GBK/GB18030 编码。')
    return []


def read_excel_targets(excel_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError('读取 Excel 需要 openpyxl，请在打包机器安装 openpyxl 后重新打包 exe。')

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return []

        fieldnames = [str(value).strip() if value is not None else '' for value in headers]
        columns = resolve_input_columns(fieldnames)
        column_indexes = {canonical: fieldnames.index(source) for canonical, source in columns.items()}

        targets = []
        for row_number, row in enumerate(rows, start=2):
            row = row or []
            target = {}
            for canonical, index in column_indexes.items():
                value = row[index] if index < len(row) else ''
                target[canonical] = '' if value is None else str(value).strip()
            if any(target.values()):
                target['_row_number'] = row_number
                targets.append(target)
        return targets
    finally:
        workbook.close()


def display_value(value):
    value = '' if value is None else str(value).strip()
    return value or '<empty>'


def parse_well_no(value):
    value = '' if value is None else str(value).strip()
    if value.endswith('.0'):
        value = value[:-2]
    return int(value)


def validate_targets(targets):
    if not targets:
        raise RuntimeError('Excel 没有可处理的数据行。')

    errors = []
    for index, target in enumerate(targets, start=2):
        row_number = target.get('_row_number') or index
        patient_given_names = (target.get('patient_given_names') or '').strip()
        patient_name = (target.get('patient_name') or '').strip()
        if not (target.get('no') or '').strip():
            errors.append('第 {} 行缺少 no'.format(row_number))
            continue
        try:
            well_no = parse_well_no(target.get('no'))
        except (TypeError, ValueError):
            errors.append('第 {} 行 no 不合法: {}'.format(row_number, display_value(target.get('no'))))
            continue
        if not patient_given_names and not patient_name:
            errors.append('第 {} 行 女方姓名 和 男方姓名 至少要填写一个'.format(row_number))
        if not 1 <= well_no <= 16:
            errors.append('第 {} 行 no 不合法: {}'.format(row_number, display_value(target.get('no'))))

    if errors:
        shown = errors[:20]
        if len(errors) > 20:
            shown.append('还有 {} 个错误未显示。'.format(len(errors) - 20))
        raise RuntimeError('Excel 数据预检失败:\n{}'.format('\n'.join(shown)))


def read_targets(input_path):
    extension = os.path.splitext(input_path)[1].lower()
    if extension in ['.xlsx', '.xlsm']:
        targets = read_excel_targets(input_path)
    elif extension == '.xls':
        raise RuntimeError('请先把 .xls 文件另存为 .xlsx 或 .csv 后再导入。')
    else:
        targets = read_csv_targets(input_path)
    validate_targets(targets)
    return targets


def check_output_dir_writable(output_dir):
    if not os.path.isdir(output_dir):
        os.makedirs(output_dir)

    fd, test_path = tempfile.mkstemp(prefix='write_test_', suffix='.tmp', dir=output_dir)
    try:
        with os.fdopen(fd, 'w', encoding='utf-8') as test_file:
            test_file.write('test')
    finally:
        if os.path.exists(test_path):
            os.remove(test_path)


def index_dishrecords(dishrecords):
    by_name = {}
    for dish in dishrecords:
        key = make_patient_key(dish.get('patient_given_names'), dish.get('patient_name'))
        by_name.setdefault(key, []).append(dish)
    return by_name


def find_matching_dishrecords(dishrecords, patient_given_names, patient_name):
    patient_given_names = normalize(patient_given_names)
    patient_name = normalize(patient_name)
    matches = []

    for dish in dishrecords:
        dish_given_names = normalize(dish.get('patient_given_names'))
        dish_patient_name = normalize(dish.get('patient_name'))
        if patient_given_names and patient_name:
            matched = dish_given_names == patient_given_names and dish_patient_name == patient_name
        elif patient_given_names:
            matched = dish_given_names == patient_given_names
        else:
            matched = dish_patient_name == patient_name
        if matched:
            matches.append(dish)
    return matches


def format_start_time(age_at_start):
    if age_at_start is None or age_at_start == '':
        return '00.00'

    try:
        seconds = int(float(age_at_start))
    except (TypeError, ValueError):
        return '00.00'
    return '{:02d}.{:02d}'.format(seconds // 3600, (seconds % 3600) // 60)


def make_video_filename(patient_given_names, patient_name, session_uuid, well_no, zid, age_at_start=None):
    name_parts = [safe_filename(value) for value in [patient_given_names, patient_name] if (value or '').strip()]
    name_parts.extend([
        safe_filename(session_uuid),
        'well{:02d}_zid{}_{}.mp4'.format(well_no, safe_filename(zid), format_start_time(age_at_start)),
    ])
    return '-'.join(name_parts)


def download_one_video(client, patient_given_names, patient_name, session_uuid, well_no, zid,
                       output_dir, overwrite, summary, progress, age_at_start=None):
    filename = make_video_filename(patient_given_names, patient_name, session_uuid, well_no, zid, age_at_start)
    output_path = os.path.join(output_dir, filename)
    try:
        result = client.download_video(session_uuid, well_no, zid, output_path, overwrite=overwrite)
        summary[result] += 1
        progress('info', '{} {}'.format('已跳过' if result == 'skipped' else '已下载', output_path))
    except Exception as ex:
        summary['failed'] += 1
        progress('error', '下载失败: {} {} session_uuid={} well{:02d} zid{}: {}'.format(
            patient_given_names, patient_name, session_uuid, well_no, zid, ex))


def download_patient_videos(client, input_path, output_dir, overwrite=False,
                            progress_callback=None, row_callback=None):
    def progress(level, message):
        if progress_callback:
            progress_callback(level, message)

    check_output_dir_writable(output_dir)

    targets = read_targets(input_path)
    summary = {'downloaded': 0, 'skipped': 0, 'not_found': 0, 'no_session': 0, 'failed': 0}
    total = len(targets)
    rows_without_session_uuid = [target for target in targets if not (target.get('session_uuid') or '').strip()]
    dishrecords = None

    if rows_without_session_uuid:
        progress('info', '正在获取 dishrecords...')
        dishrecords = client.get_dishrecords()

    for index, target in enumerate(targets, start=1):
        if row_callback:
            row_callback(index - 1, total)

        patient_given_names = target.get('patient_given_names')
        patient_name = target.get('patient_name')
        row_number = target.get('_row_number') or index + 1
        well_no = parse_well_no(target.get('no'))
        session_uuid = (target.get('session_uuid') or '').strip()

        if session_uuid:
            direct_sessions = client.get_sessionrecord(session_uuid)
            if len(direct_sessions) > 1:
                summary['failed'] += 1
                progress('error', 'Excel 第 {} 行：session_uuid 查询到重复的数据，为避免配对错误已跳过该患者: session_uuid={}, session 数量={}'.format(
                    row_number, session_uuid, len(direct_sessions)))
                if row_callback:
                    row_callback(index, total)
                continue
            if not direct_sessions:
                summary['no_session'] += 1
                progress('warning', 'Excel 第 {} 行：当前服务器未找到 session_uuid 对应的数据，已跳过，可能不是这个 IP 的数据: session_uuid={}'.format(
                    row_number, session_uuid))
                if row_callback:
                    row_callback(index, total)
                continue
            age_at_start = direct_sessions[0].get('age_at_start')
            for zid in VIDEO_ZIDS:
                download_one_video(client, patient_given_names, patient_name, session_uuid, well_no, zid,
                                   output_dir, overwrite, summary, progress, age_at_start=age_at_start)
            if row_callback:
                row_callback(index, total)
            continue

        matches = find_matching_dishrecords(dishrecords, patient_given_names, patient_name)

        if not matches:
            summary['not_found'] += 1
            progress('warning', 'Excel 第 {} 行：当前服务器未找到 dish，已跳过: {} {} well{:02d}'.format(
                row_number, patient_given_names, patient_name, well_no))
            if row_callback:
                row_callback(index, total)
            continue
        if len(matches) > 1:
            summary['failed'] += 1
            progress('error', 'Excel 第 {} 行：查询到重复的数据，为避免配对错误已跳过该患者。可以在 Excel 添加一列 session_uuid，用来匹配正确的数据: {} {} well{:02d}, dish 数量={}'.format(
                row_number, patient_given_names, patient_name, well_no, len(matches)))
            if row_callback:
                row_callback(index, total)
            continue

        for dish in matches:
            dish_uuid = dish.get('dish_uuid')
            sessions = client.get_sessionrecords(dish_uuid)
            if not sessions:
                summary['no_session'] += 1
                progress('warning', 'Excel 第 {} 行：当前服务器未找到 session，已跳过: {} {} dish_uuid={}'.format(
                    row_number, patient_given_names, patient_name, dish_uuid))
                continue
            if len(sessions) > 1:
                summary['failed'] += 1
                progress('error', 'Excel 第 {} 行：查询到重复的数据，为避免配对错误已跳过该患者。可以在 Excel 添加一列 session_uuid，用来匹配正确的数据: {} {} dish_uuid={}, session 数量={}'.format(
                    row_number, patient_given_names, patient_name, dish_uuid, len(sessions)))
                if row_callback:
                    row_callback(index, total)
                continue

            for session in sessions:
                session_uuid = session.get('session_uuid')
                if not session_uuid:
                    summary['no_session'] += 1
                    progress('warning', 'Excel 第 {} 行：sessionrecord 缺少 session_uuid，已跳过: {} {} dish_uuid={}'.format(
                        row_number, patient_given_names, patient_name, dish_uuid))
                    continue
                age_at_start = session.get('age_at_start')
                for zid in VIDEO_ZIDS:
                    download_one_video(client, patient_given_names, patient_name, session_uuid, well_no, zid,
                                       output_dir, overwrite, summary, progress, age_at_start=age_at_start)

        if row_callback:
            row_callback(index, total)

    return summary


class DownloadPatientVideosApp:

    def __init__(self, root):
        self.root = root
        self.root.title('患者视频下载工具')
        self.root.geometry('1200x650')
        self.events = queue.Queue()
        self.worker = None

        self.input_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.abspath('videos'))
        self.server = tk.StringVar(value='https://10.100.22.10')
        self.username = tk.StringVar(value='admin')
        self.password = tk.StringVar(value='admin')
        self.overwrite = tk.BooleanVar(value=False)
        self.progress_text = tk.StringVar(value='请选择 Excel 文件并填写服务器信息')

        self._build_ui()
        self.root.after(100, self._process_events)

    def _build_ui(self):
        frame = ttk.Frame(self.root, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(9, weight=1)


        help_text = (
            "使用说明：\n"
            "1. Excel 必须包含表头 [女方姓名]、[男方姓名]、[no], 其中 [session_uuid] 为可选表头。\n"
            "2. 女方姓名]、[男方姓名] 需与 GCA 患者页面的 [名字]、[姓氏] 表头数据保持一致。\n"
            "3. [no] 为必填项，表示位孔编号。\n"
            "4. 如程序查询到重复数据，将自动跳过该行并标红提示。请在 Excel 中补充[session_uuid]列后重新点击预检并下载，以避免匹配到错误数据。"
        )

        ttk.Label(frame, text=help_text, foreground='#004080', wraplength=940, justify=tk.LEFT).grid(
            row=0, column=0, columnspan=3, sticky=tk.EW, pady=(0, 10))

        ttk.Label(frame, text='Excel 文件').grid(row=1, column=0, sticky=tk.W, pady=4)
        ttk.Entry(frame, textvariable=self.input_path).grid(row=1, column=1, sticky=tk.EW, pady=4)
        ttk.Button(frame, text='选择...', command=self._browse_input).grid(row=1, column=2, padx=(6, 0), pady=4)

        ttk.Label(frame, text='保存目录').grid(row=2, column=0, sticky=tk.W, pady=4)
        ttk.Entry(frame, textvariable=self.output_dir).grid(row=2, column=1, sticky=tk.EW, pady=4)
        ttk.Button(frame, text='选择...', command=self._browse_output_dir).grid(row=2, column=2, padx=(6, 0), pady=4)

        ttk.Label(frame, text='服务器 IP/地址').grid(row=3, column=0, sticky=tk.W, pady=4)
        server_frame = ttk.Frame(frame)
        server_frame.grid(row=3, column=1, columnspan=2, sticky=tk.EW, pady=4)
        server_frame.columnconfigure(1, weight=1)
        server_frame.columnconfigure(3, weight=1)
        server_frame.columnconfigure(5, weight=1)
        ttk.Entry(server_frame, textvariable=self.server).grid(row=0, column=1, sticky=tk.EW)
        ttk.Label(server_frame, text=' 登录用户名 ').grid(row=0, column=2)
        ttk.Entry(server_frame, textvariable=self.username).grid(row=0, column=3, sticky=tk.EW)
        ttk.Label(server_frame, text=' 登录密码 ').grid(row=0, column=4)
        ttk.Entry(server_frame, textvariable=self.password).grid(row=0, column=5, sticky=tk.EW)

        options = ttk.Frame(frame)
        options.grid(row=4, column=1, columnspan=2, sticky=tk.W, pady=4)
        ttk.Checkbutton(options, text='覆盖已存在视频', variable=self.overwrite).pack(side=tk.LEFT)

        self.start_button = ttk.Button(frame, text='预检并开始下载', command=self._start_download)
        self.start_button.grid(row=5, column=0, columnspan=3, sticky=tk.EW, pady=(12, 6))

        self.progress_bar = ttk.Progressbar(frame, mode='determinate')
        self.progress_bar.grid(row=6, column=0, columnspan=3, sticky=tk.EW)
        ttk.Label(frame, textvariable=self.progress_text).grid(row=7, column=0, columnspan=3, sticky=tk.NW, pady=(6, 0))

        self.error_text = tk.StringVar()
        self.error_label = ttk.Label(frame, textvariable=self.error_text, foreground='red', wraplength=900)
        self.error_label.grid(row=8, column=0, columnspan=3, sticky=tk.EW, pady=(6, 0))

        log_frame = ttk.Frame(frame)
        log_frame.grid(row=9, column=0, columnspan=3, sticky=tk.NSEW, pady=(12, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_text = tk.Text(log_frame, height=16, wrap=tk.WORD)
        self.log_text.tag_configure('ERROR', foreground='red')
        self.log_text.tag_configure('WARNING', foreground='red')
        self.log_text.tag_configure('INFO', foreground='black')
        self.log_text.grid(row=0, column=0, sticky=tk.NSEW)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        scrollbar.grid(row=0, column=1, sticky=tk.NS)
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def _browse_input(self):
        filename = filedialog.askopenfilename(
            title='选择患者 Excel/CSV 文件',
            filetypes=[('Excel/CSV files', '*.xlsx *.xlsm *.csv'), ('Excel files', '*.xlsx *.xlsm'),
                       ('CSV files', '*.csv'), ('All files', '*.*')])
        if filename:
            self.input_path.set(filename)

    def _browse_output_dir(self):
        dirname = filedialog.askdirectory(title='选择视频保存目录')
        if dirname:
            self.output_dir.set(dirname)

    def _validate_inputs(self):
        input_path = self.input_path.get().strip()
        output_dir = self.output_dir.get().strip()
        server = self.server.get().strip()
        username = self.username.get().strip()
        password = self.password.get()

        if not input_path:
            raise RuntimeError('请选择 Excel 或 CSV 文件。')
        if not os.path.isfile(input_path):
            raise RuntimeError('文件不存在: {}'.format(input_path))
        if not output_dir:
            raise RuntimeError('请选择视频保存目录。')
        if not (server.startswith('https://') or server.startswith('http://')):
            raise RuntimeError('服务器地址必须以 http:// 或 https:// 开头。')
        if not username:
            raise RuntimeError('请输入登录用户名。')
        if not password:
            raise RuntimeError('请输入登录密码。')

        check_output_dir_writable(output_dir)
        targets = read_targets(input_path)
        self._append_log('INFO', '预检通过：Excel 表头和数据正常，共 {} 行。'.format(len(targets)))
        self._append_log('INFO', '预检通过：视频保存目录可以写入。')
        self._append_log('INFO', '正在预检 IP、用户名和密码...')
        client = CandlePublicClient(server, username, password)
        self._append_log('INFO', '预检通过：IP 可以连接，用户名和密码可以获取 auth。')
        return input_path, output_dir, server, client

    def _start_download(self):
        if self.worker and self.worker.is_alive():
            return

        self._clear_log()
        self.error_text.set('')
        self.progress_bar['value'] = 0
        self.progress_text.set('正在预检...')

        try:
            input_path, output_dir, server, client = self._validate_inputs()
        except Exception as ex:
            self.progress_text.set('预检失败')
            self._show_error('预检失败: {}'.format(ex))
            return

        self.progress_text.set('开始下载...')
        self.start_button.configure(state=tk.DISABLED)

        args = {
            'input_path': input_path,
            'output_dir': output_dir,
            'server': server,
            'overwrite': self.overwrite.get(),
            'client': client,
        }
        self.worker = threading.Thread(target=self._download_worker, kwargs=args, daemon=True)
        self.worker.start()

    def _download_worker(self, input_path, output_dir, server, overwrite, client):
        try:
            self.events.put(('log', 'INFO', '开始下载，服务器 {}'.format(server)))

            def progress_callback(level, message):
                self.events.put(('log', level.upper(), message))

            def row_callback(current, total):
                self.events.put(('progress', current, total))

            summary = download_patient_videos(
                client,
                input_path,
                output_dir,
                overwrite=overwrite,
                progress_callback=progress_callback,
                row_callback=row_callback,
            )
            self.events.put(('done', summary))
        except Exception as ex:
            self.events.put(('failed', str(ex)))

    def _process_events(self):
        try:
            while True:
                event = self.events.get_nowait()
                event_type = event[0]
                if event_type == 'log':
                    self._append_log(event[1], event[2])
                elif event_type == 'progress':
                    current, total = event[1], event[2]
                    self.progress_bar['maximum'] = total or 1
                    self.progress_bar['value'] = current
                    self.progress_text.set('已处理 {} / {}'.format(current, total))
                elif event_type == 'done':
                    self._finish(event[1])
                elif event_type == 'failed':
                    self._fail(event[1])
        except queue.Empty:
            pass
        self.root.after(100, self._process_events)

    def _append_log(self, level, message):
        tag = level if level in ['ERROR', 'WARNING', 'INFO'] else 'INFO'
        self.log_text.insert(tk.END, '[{}] {}\n'.format(level, message), tag)
        self.log_text.see(tk.END)

    def _clear_log(self):
        self.log_text.delete('1.0', tk.END)

    def _finish(self, summary):
        self.start_button.configure(state=tk.NORMAL)
        self.error_text.set('')
        self.progress_text.set('下载完成')
        text = json.dumps(summary, sort_keys=True, ensure_ascii=False)
        self._append_log('INFO', 'Summary: {}'.format(text))

    def _show_error(self, message):
        self.error_text.set(message)
        self._append_log('ERROR', message)

    def _fail(self, message):
        self.start_button.configure(state=tk.NORMAL)
        self.progress_text.set('下载失败')
        self._show_error(message)


def main():
    requests.packages.urllib3.disable_warnings()
    root = tk.Tk()
    app = DownloadPatientVideosApp(root)
    root.mainloop()
    return app


if __name__ == '__main__':
    main()
